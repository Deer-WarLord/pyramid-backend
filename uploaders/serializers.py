from rest_framework import serializers
from openpyxl import load_workbook
from uploaders.models import *
from rest_framework.fields import FileField
from io import BytesIO
from django.utils.translation import ugettext_lazy as _
from zipfile import ZipFile, is_zipfile
import json
import unicodecsv as csv
from os.path import splitext


class JsonFileField(FileField):
    default_error_messages = {
        'required': _('No file was submitted.'),
        'invalid': _('The submitted data was not a file. Check the encoding type on the form.'),
        'no_name': _('No filename could be determined.'),
        'empty': _('The submitted file is empty.'),
        'max_length': _('Ensure this filename has at most {max_length} characters (it has {length}).'),
        'json': _('The submitted file has invalid format. Should be json-format.'),
    }

    def to_internal_value(self, data):
        super(JsonFileField, self).to_internal_value(data)

        if is_zipfile(data):
            with ZipFile(data) as zf:
                raw_data = zf.read(splitext(data.name)[0])
        else:
            data.seek(0)
            raw_data = data.read()

        try:
            data.json = json.loads(raw_data)
        except ValueError:
            try:
                data.json = json.loads(raw_data, encoding='cp1251')
            except ValueError:
                try:
                    lines = raw_data.splitlines()
                    dialect = csv.Sniffer().sniff(lines[0], [',', ';', '\t'])
                    data.json = [item for item in csv.DictReader(lines, dialect=dialect)]
                except (ValueError, csv.Error):
                    self.fail('json')
        return data


class XlsxFileField(FileField):
    default_error_messages = {
        'required': _('No file was submitted.'),
        'invalid': _('The submitted data was not a file. Check the encoding type on the form.'),
        'no_name': _('No filename could be determined.'),
        'empty': _('The submitted file is empty.'),
        'max_length': _('Ensure this filename has at most {max_length} characters (it has {length}).'),
        'xlsx': _('The submitted file has invalid format. Should be xlsx-format.'),
    }

    def to_internal_value(self, data):
        super(XlsxFileField, self).to_internal_value(data)

        workbook = load_workbook(filename=BytesIO(data.read()))
        worksheet = workbook[workbook.sheetnames[0]]
        iter_rows = worksheet.iter_rows()
        next(iter_rows)
        results = []
        fields = ("id", "key_word", "title", "posted_date", "posted_time", "end_time", "publication", "rat", "shr")
        for row in iter_rows:
            publication = dict(zip(fields, [cell.value for cell in row]))
            del publication["id"]
            item = {
                "rat": publication.pop("rat"),
                "shr": publication.pop("shr"),
                "publication": publication
            }
            results.append(item)

        data.json = results

        return data


class UploadedInfoSerializer(serializers.ModelSerializer):
    class Meta:
        model = UploadedInfo
        fields = ('id', 'provider', 'title', 'file', 'created_date')

    provider = serializers.ReadOnlyField(source='provider.title')
    file = JsonFileField()
    created_date = serializers.ReadOnlyField()


class UploadedInfoTvSerializer(serializers.ModelSerializer):
    class Meta:
        model = UploadedInfo
        fields = ('id', 'provider', 'title', 'file', 'created_date')

    file = XlsxFileField()
    created_date = serializers.ReadOnlyField()


class UploadedInfoSerializerOwner(serializers.ModelSerializer):
    class Meta:
        model = UploadedInfo
        fields = ('id', 'provider', 'title', 'file', 'created_date')

    file = JsonFileField()
    created_date = serializers.ReadOnlyField()


class AsyncUploadedInfoSerializer(serializers.Serializer):
    title = serializers.CharField(max_length=256)
    provider = serializers.ReadOnlyField()
    url = serializers.URLField()


class AsyncUploadedInfoSerializerOwner(serializers.Serializer):
    title = serializers.CharField(max_length=256)
    provider = serializers.IntegerField()
    url = serializers.URLField()
    send_to_provider = serializers.CharField(max_length=5, required=False, default="off")